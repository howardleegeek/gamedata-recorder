#!/usr/bin/env python3
"""Generate gameinfo.xlsx for a recording session per PRD §3.3.

PRD §3.3 mandates a SINGLE sheet with 14 specific fields (in order):
  game_name, game_version, platform, scene_name, weather, time_of_day,
  character_name, character_class, operator_id, recording_date,
  total_frames, video_duration_sec, route_type, notes

Sources:
  * metadata.json (recorder-written)        — session_id, duration, hardware
  * frames.jsonl (recorder-written)         — total_frames count
  * env vars (operator-supplied)            — operator_id, character_*, notes
  * defaults                                — game_name="Minecraft", platform="Java Edition"

Fields needing mc-mod IPC (weather, time_of_day, scene_name) currently use
sensible defaults; rc17.4+ will pipe real values from mc-mod GameStateSample.

rc17.3.1 (Stream BJ-rewrite, Howard "必须解决" 2026-05-12): replaces the
4-sheet placeholder workbook with the canonical PRD single-sheet 14-field
schema. Customer rejects 4-sheet variant.
"""

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def load_metadata(session_dir: Path) -> Dict[str, Any]:
    """Load metadata from metadata.json or systeminfo.json."""
    metadata_path = session_dir / "metadata.json"
    if not metadata_path.exists():
        metadata_path = session_dir / "systeminfo.json"
    if not metadata_path.exists():
        raise FileNotFoundError(f"No metadata file found in {session_dir}")
    with open(metadata_path, 'r') as f:
        return json.load(f)


def count_frames(session_dir: Path) -> int:
    """Count entries in frames.jsonl.

    Bug-fix 2026-05-15: use ``with`` context manager so the file handle is
    closed deterministically even if the generator is interrupted partway.
    """
    p = session_dir / "frames.jsonl"
    if not p.is_file():
        return 0
    with p.open('r', encoding='utf-8') as fh:
        return sum(1 for line in fh if line.strip())


def derive_recording_date(metadata: Dict[str, Any]) -> str:
    """Derive recording_date YYYY-MM-DD from metadata.start_timestamp.

    Bug-fix 2026-05-15: ``datetime.utcnow()`` is deprecated in Python 3.12+
    (and slated for removal); use timezone-aware ``datetime.now(timezone.utc)``.
    Same applies to ``datetime.fromtimestamp(ts)`` — pass ``tz=timezone.utc``
    so the returned object isn't a naive local-time value.
    """
    ts = metadata.get("start_timestamp")
    if ts is None:
        return datetime.now(timezone.utc).strftime("%Y-%m-%d")
    try:
        return datetime.fromtimestamp(float(ts), tz=timezone.utc).strftime("%Y-%m-%d")
    except (ValueError, TypeError, OSError):
        return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def derive_duration(metadata: Dict[str, Any]) -> float:
    """Get video_duration_sec from metadata."""
    if "duration" in metadata:
        return float(metadata["duration"])
    start = metadata.get("start_timestamp")
    end = metadata.get("end_timestamp")
    if start is not None and end is not None:
        return float(end) - float(start)
    return 0.0


def _detect_mc_username() -> Optional[str]:
    """RBGA-C3: try to read the real MC username from launcher_profiles.json
    so character_name in gameinfo.xlsx isn't the hardcoded "DataPilot".

    Returns None if not findable — caller falls back to default. We check
    the standard MC launcher paths in priority order:
      - %APPDATA%/.minecraft/launcher_profiles.json (Windows)
      - ~/Library/Application Support/minecraft/launcher_profiles.json (mac)
      - ~/.minecraft/launcher_profiles.json (linux)
      - bundled mc-instance under the recorder install (rc19.0.x bundles MC)
    """
    candidates = []
    appdata = os.environ.get("APPDATA")
    if appdata:
        candidates.append(Path(appdata) / ".minecraft" / "launcher_profiles.json")
    home = Path.home()
    candidates += [
        home / "Library" / "Application Support" / "minecraft" / "launcher_profiles.json",
        home / ".minecraft" / "launcher_profiles.json",
        home / "AppData" / "Local" / "GameData Recorder" / "mc-instance" / "launcher_profiles.json",
    ]
    for p in candidates:
        try:
            if not p.is_file():
                continue
            data = json.loads(p.read_text())
            # MC launcher_profiles.json has a `selectedUser` key with UUID;
            # the actual username lives in `authenticationDatabase.<uuid>.username`
            # in older versions, or directly visible in `profiles.<id>.name`.
            # Try multiple shapes for resilience across launcher versions.
            sel = data.get("selectedUser") or {}
            if isinstance(sel, dict):
                uuid_ = sel.get("account")
                auth = data.get("authenticationDatabase", {})
                if uuid_ and uuid_ in auth:
                    name = auth[uuid_].get("username") or auth[uuid_].get("displayName")
                    if name:
                        return str(name)
            # Newer launchers: try first profile's name
            for prof_id, prof in (data.get("profiles") or {}).items():
                name = prof.get("name") or prof.get("lastVersionId")
                if name and name not in ("(Default)", "latest-release"):
                    return str(name)
        except (json.JSONDecodeError, OSError, KeyError):
            continue
    return None


def _parse_route_type(raw: str) -> int:
    """Parse + clamp route_type to PRD-allowed values {1,2,3}.

    Returns 1 on any error (non-numeric, out-of-range, None). Emits a stderr
    warning so the bad config is visible, but does NOT raise — the recording
    is more important than perfect provenance, and D4 audit will catch it.
    """
    try:
        rt = int(raw)
    except (TypeError, ValueError):
        print(f"[gameinfo] WARN: OYSTER_ROUTE_TYPE={raw!r} not an int, using 1",
              file=sys.stderr)
        return 1
    if rt not in (1, 2, 3):
        print(f"[gameinfo] WARN: OYSTER_ROUTE_TYPE={rt} not in {{1,2,3}}, using 1",
              file=sys.stderr)
        return 1
    return rt


# MECE I3 — cyclic route_type counter file.
# Persists across recordings so operator gets natural 1→2→3→1 rotation
# without needing to set OYSTER_ROUTE_TYPE manually each session. Explicit
# env override still wins (see _resolve_route_type below).
_ROUTE_COUNTER_PATH = Path.home() / ".oyster-route-counter"


def _next_cyclic_route_type() -> int:
    """Read + increment + write the persistent route counter.

    Cycles 1 → 2 → 3 → 1 → ... across recordings. File is a single-byte
    text file (digit '1'/'2'/'3'). Corrupt / missing → start at 1.

    NOT atomic across concurrent recorders (unlikely on a single rig);
    if that ever matters add fcntl.flock around the read-modify-write.
    """
    try:
        raw = _ROUTE_COUNTER_PATH.read_text().strip()
        prev = int(raw)
        if prev not in (1, 2, 3):
            prev = 0  # treat as "before first" — first call returns 1
    except (OSError, ValueError):
        prev = 0
    nxt = (prev % 3) + 1  # 0→1, 1→2, 2→3, 3→1
    try:
        _ROUTE_COUNTER_PATH.write_text(str(nxt))
    except OSError as e:
        print(f"[gameinfo] WARN: could not persist route counter at "
              f"{_ROUTE_COUNTER_PATH}: {e}", file=sys.stderr)
    return nxt


def _resolve_route_type() -> int:
    """MECE I3 — pick route_type with priority:
      1. OYSTER_ROUTE_TYPE env (explicit override per recording)
      2. Cyclic counter at ~/.oyster-route-counter (default behavior)
      3. Fallback 1 if everything fails (matches old behavior)
    """
    env_raw = os.environ.get("OYSTER_ROUTE_TYPE")
    if env_raw is not None and env_raw != "":
        # Explicit override — don't advance the cyclic counter (else
        # setting OYSTER_ROUTE_TYPE=2 once would shift cyclic too).
        return _parse_route_type(env_raw)
    return _next_cyclic_route_type()


# MECE G15 — strict operator_id with persistent config + loud sentinel.
_OPERATOR_CONFIG_PATH = Path.home() / ".oyster-operator.json"
_OPERATOR_MISSING_SENTINEL = "operator-MISSING-CONFIG"


def _resolve_operator_id() -> str:
    """MECE G15 / RBGA-C2 — pick operator_id with priority:
      1. OYSTER_OPERATOR_ID env (per-recording override)
      2. ~/.oyster-operator.json {"operator_id": "..."} (persistent config)
      3. Loud sentinel ``operator-MISSING-CONFIG`` so the bad provenance
         is impossible to ignore — buyer pipeline lint flags it, scorecard
         shows red, and grep finds every leaked session.

    Previously defaulted to the literal string "vendor-001-op-A" which
    silently contaminated every untraceable session with what looked
    like a real ID. The sentinel removes ambiguity.
    """
    env_raw = os.environ.get("OYSTER_OPERATOR_ID")
    if env_raw is not None and env_raw.strip() != "":
        return env_raw.strip()
    try:
        if _OPERATOR_CONFIG_PATH.is_file():
            data = json.loads(_OPERATOR_CONFIG_PATH.read_text(encoding="utf-8"))
            op = data.get("operator_id") if isinstance(data, dict) else None
            if isinstance(op, str) and op.strip():
                return op.strip()
    except (json.JSONDecodeError, OSError) as e:
        print(f"[gameinfo] WARN: could not read {_OPERATOR_CONFIG_PATH}: {e}",
              file=sys.stderr)
    print(f"[gameinfo] ERROR: operator_id NOT configured "
          f"(neither OYSTER_OPERATOR_ID env nor {_OPERATOR_CONFIG_PATH}). "
          f"Writing sentinel {_OPERATOR_MISSING_SENTINEL!r}. "
          f"Set OYSTER_OPERATOR_ID or run: echo '{{\"operator_id\":\"tester-N\"}}' "
          f"> {_OPERATOR_CONFIG_PATH}", file=sys.stderr)
    return _OPERATOR_MISSING_SENTINEL


# PRD §3.3 — 14 fields in this exact order
PRD_FIELD_ORDER = [
    "game_name",
    "game_version",
    "platform",
    "scene_name",
    "weather",
    "time_of_day",
    "character_name",
    "character_class",
    "operator_id",
    "recording_date",
    "total_frames",
    "video_duration_sec",
    "route_type",
    "notes",
]


def build_row(metadata: Dict[str, Any], session_dir: Path) -> Dict[str, Any]:
    """Build the 14-field row per PRD §3.3.

    Operator-configurable fields are pulled from env vars; defaults are
    used when the env is unset. Defaults pass the lint v3 schema check
    but should be overridden via env for each session in production.
    """
    return {
        "game_name": os.environ.get("OYSTER_GAME_NAME", "Minecraft"),
        # rc17.3.1: prefer env override; recorder-side detection of MC version
        # is wired in rc17.4 via mc-mod IPC.
        "game_version": os.environ.get("OYSTER_GAME_VERSION", "1.21.4"),
        "platform": os.environ.get("OYSTER_PLATFORM", "Java Edition"),
        # scene_name / weather / time_of_day need mc-mod IPC (deferred to rc17.4).
        # rc17.3.1 fallback: configurable via env, defaults pass schema check.
        "scene_name": os.environ.get("OYSTER_SCENE_NAME", "flat-overworld"),
        "weather": os.environ.get("OYSTER_WEATHER", "clear"),
        "time_of_day": os.environ.get("OYSTER_TIME_OF_DAY", "day"),
        # Character + operator metadata: operator-supplied via env / launcher
        # form (launcher form UI is rc17.4). rc19.0.5 (RBGA-C3): fallback to
        # MC's launcher_profiles.json so character_name = real MC username
        # rather than placeholder. Env var still wins for explicit override.
        "character_name": os.environ.get("OYSTER_CHARACTER_NAME") or _detect_mc_username() or "DataPilot",
        "character_class": os.environ.get("OYSTER_CHARACTER_CLASS", "survival"),
        # MECE G15 — strict operator_id resolution (env → config → sentinel).
        "operator_id": _resolve_operator_id(),
        "recording_date": derive_recording_date(metadata),
        "total_frames": count_frames(session_dir),
        "video_duration_sec": round(derive_duration(metadata), 2),
        # route_type ∈ {1,2,3}; PRD-defined route classes.
        # MECE I3 (2026-05-15): cyclic counter persisted at ~/.oyster-route-counter
        # so operator gets natural 1→2→3→1 rotation across recordings.
        # OYSTER_ROUTE_TYPE env still overrides for single-session pinning.
        # Bad/missing input falls back to 1 with stderr warning (does NOT
        # crash the pipeline — recording is more important than provenance).
        "route_type": _resolve_route_type(),
        "notes": os.environ.get("OYSTER_NOTES", ""),
    }


def create_workbook(session_dir: Path, output_path: Optional[Path] = None) -> int:
    """Create the gameinfo.xlsx workbook per PRD §3.3."""
    session_dir = Path(session_dir)
    metadata = load_metadata(session_dir)
    row = build_row(metadata, session_dir)

    wb = Workbook()
    ws = wb.active
    ws.title = "gameinfo"

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")

    # Header row in PRD order
    for col_idx, field in enumerate(PRD_FIELD_ORDER, start=1):
        c = ws.cell(row=1, column=col_idx, value=field)
        c.font = header_font
        c.alignment = header_alignment

    # Single data row
    for col_idx, field in enumerate(PRD_FIELD_ORDER, start=1):
        ws.cell(row=2, column=col_idx, value=row[field])

    # Column widths
    for col_idx, field in enumerate(PRD_FIELD_ORDER, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(len(str(row[field])) + 2, len(field) + 2, 12)

    out = Path(output_path) if output_path else (session_dir / "gameinfo.xlsx")
    wb.save(out)
    print(f"[gameinfo] wrote {out} with {len(PRD_FIELD_ORDER)} PRD §3.3 fields")
    return len(PRD_FIELD_ORDER)


def main():
    p = argparse.ArgumentParser(description="Generate gameinfo.xlsx per PRD §3.3")
    p.add_argument("session_dir", type=Path, help="Recording session directory")
    p.add_argument("--output", type=Path, default=None,
                   help="Output path (default: session_dir/gameinfo.xlsx)")
    args = p.parse_args()
    create_workbook(args.session_dir, args.output)


if __name__ == "__main__":
    main()
